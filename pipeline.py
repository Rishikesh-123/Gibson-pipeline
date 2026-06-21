"""
Gibson Bank Lending Signal Pipeline
=====================================
1. Pulls 10-K filings from SEC EDGAR (free, no API key needed)
2. Extracts debt notes using edgartools built-in sections
3. Sends to Claude AI to extract structured data
4. Saves everything to a formatted Excel file
"""

import os
import sys
import time
import json
import re
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    from edgar import Company, set_identity
except ImportError:
    print("ERROR: Run this first:  pip install edgartools")
    sys.exit(1)

from config import ANTHROPIC_API_KEY, EDGAR_IDENTITY, TICKERS

# ── Setup ──────────────────────────────────────────────────────
set_identity(EDGAR_IDENTITY)
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
os.makedirs("output", exist_ok=True)

# ── Colors for Excel ───────────────────────────────────────────
LENDER_COLORS = {
    "Traditional commercial bank":    "C6EFCE",
    "Specialty/venture debt":         "FFEB9C",
    "Royalty/alternative":            "FCE4D6",
    "Related-party":                  "F4B8D1",
    "Institutional convertible/ELOC": "D9D9D9",
    "None":                           "FFFFFF",
    "Unknown":                        "F2F2F2",
}

# ═══════════════════════════════════════════════════════════════
# STEP 1 — PULL FILING FROM EDGAR
# ═══════════════════════════════════════════════════════════════

def get_debt_section(ticker):
    result = {
        "ticker":                    ticker,
        "filing_date":               None,
        "period":                    None,
        "accession":                 None,
        "debt_note":                 None,
        "balance_sheet_liabilities": None,
        "mda_liquidity":             None,
        "error":                     None,
    }

    try:
        company = Company(ticker)
        filings = company.get_filings(form="10-K")
        latest  = filings.latest(1)

        result["period"]      = str(getattr(latest, 'period_of_report', 'Unknown'))
        result["accession"]   = str(getattr(latest, 'accession_no', 'Unknown'))
        result["filing_date"] = str(getattr(latest, 'filed', 'Unknown'))

        filing = latest.obj()

        # ── DEBT NOTE: search notes by title ──────────────────────
        debt_keywords = [
            'debt', 'credit facility', 'borrowing', 'notes payable',
            'long-term debt', 'financing', 'loan', 'revolving'
        ]
        try:
            notes      = filing.notes
            debt_notes = [
                n for n in notes
                if any(kw in n.title.lower() for kw in debt_keywords)
            ]
            if debt_notes:
                result["debt_note"] = str(debt_notes[0].text)[:6000]
        except Exception as e:
            result["error"] = f"Notes error: {str(e)[:100]}"

        # ── BALANCE SHEET ──────────────────────────────────────────
        try:
            bs = filing.balance_sheet
            if bs is not None:
                result["balance_sheet_liabilities"] = str(bs)[:2000]
        except Exception:
            pass

        # ── MD&A LIQUIDITY ─────────────────────────────────────────
        try:
            mda      = filing.management_discussion
            if mda is not None:
                mda_text  = str(mda)[:8000]
                liq_match = re.search(
                    r'(?i)(liquidity[\s\S]{0,4000})', mda_text
                )
                result["mda_liquidity"] = (
                    liq_match.group(0)[:3000] if liq_match else mda_text[:3000]
                )
        except Exception:
            pass

        if not result["debt_note"]:
            result["error"] = "No debt note found — company may have no debt"

    except Exception as e:
        result["error"] = f"EDGAR error: {str(e)[:200]}"

    return result


# ═══════════════════════════════════════════════════════════════
# STEP 2 — SEND TO CLAUDE FOR EXTRACTION
# ═══════════════════════════════════════════════════════════════

PROMPT = """You are a financial research assistant analysing SEC 10-K filings.

Extract structured debt information from the filing text below and return ONLY a JSON object — no explanation, no markdown, just the raw JSON.

LENDER TYPE RULES:
- "Traditional commercial bank" = named depository bank (JPMorgan, Wells Fargo, Truist, BofA, Regions, PNC, KeyBank, Alerus, Old National, Pinnacle, Western Alliance, any "National Bank" or "State Bank")
- "Specialty/venture debt" = specialty lenders with rate >10% or warrants attached (K2 HealthVentures, Hercules, Silicon Valley Bank, Horizon Technology)
- "Royalty/alternative" = royalty purchase agreements, revenue-based repayment
- "Related-party" = CEO, founders, controlling shareholder, officers/directors as lenders
- "Institutional convertible/ELOC" = unnamed institutional investors, convertible notes with OID, equity lines of credit
- "None" = zero debt on the balance sheet

TYPOLOGY RULES:
- Type 1: Traditional bank, stable
- Type 2: No debt at all
- Type 3: Specialty/venture debt
- Type 4: Traditional bank, evolving/amended relationship
- Type 5: Large-cap contrast (exclude if >$300M assets)
- Type 6: Related-party web, bank wound down
- Type 7: Multi-subsidiary portfolio of bank relationships
- Type 8: Distressed/reverse-merger, entirely related-party or institutional, going concern

Return exactly this JSON structure:
{
  "has_debt": true or false,
  "lender_name": "Name of lender or None",
  "lender_type": "one of the six categories above",
  "facility_amount_m": number or null,
  "amount_drawn_m": number or null,
  "interest_rate": "e.g. SOFR+3.5% or 8.5% or null",
  "maturity": "year or date as string or null",
  "has_covenants": true or false or null,
  "covenant_summary": "brief description or null",
  "going_concern": true or false,
  "scrapability": "Excellent or Good or Moderate or Poor",
  "filing_note_location": "e.g. Note 5, Long-Term Debt",
  "typology": "Type 1 through Type 8 or Unknown",
  "key_finding": "one sentence plain English summary"
}

FILING TEXT:
{filing_text}
"""


def extract_with_claude(filing_data):
    sections = []

    if filing_data.get("balance_sheet_liabilities"):
        sections.append(
            "=== BALANCE SHEET ===\n" + filing_data["balance_sheet_liabilities"]
        )
    if filing_data.get("debt_note"):
        sections.append(
            "=== DEBT FOOTNOTE ===\n" + filing_data["debt_note"]
        )
    if filing_data.get("mda_liquidity"):
        sections.append(
            "=== MD&A LIQUIDITY ===\n" + filing_data["mda_liquidity"]
        )

    if not sections:
        return {
            "has_debt":           None,
            "lender_name":        "Could not extract",
            "lender_type":        "Unknown",
            "facility_amount_m":  None,
            "amount_drawn_m":     None,
            "interest_rate":      None,
            "maturity":           None,
            "has_covenants":      None,
            "covenant_summary":   None,
            "going_concern":      False,
            "scrapability":       "Poor",
            "filing_note_location": "N/A",
            "typology":           "Unknown",
            "key_finding":        f"No text extracted. Error: {filing_data.get('error', 'Unknown')}",
        }

    filing_text = "\n\n".join(sections)[:12000]
    prompt      = PROMPT.replace("{filing_text}", filing_text)

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response.content[0].text.strip()
        raw = raw.replace("```json", "").replace("```", "").strip()
        return json.loads(raw)

    except Exception as e:
        return {
            "has_debt":           None,
            "lender_name":        "Error",
            "lender_type":        "Unknown",
            "facility_amount_m":  None,
            "amount_drawn_m":     None,
            "interest_rate":      None,
            "maturity":           None,
            "has_covenants":      None,
            "covenant_summary":   None,
            "going_concern":      False,
            "scrapability":       "Poor",
            "filing_note_location": "N/A",
            "typology":           "Unknown",
            "key_finding":        f"Claude error: {str(e)[:150]}",
        }


# ═══════════════════════════════════════════════════════════════
# STEP 3 — SAVE TO EXCEL
# ═══════════════════════════════════════════════════════════════

def build_excel(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.freeze_panes = "A3"

    NAVY   = "1F4E78"
    thin   = Side(style="thin", color="CCCCCC")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def fill(c): return PatternFill("solid", fgColor=c)
    def hf():    return Font(name="Arial", bold=True, size=10, color="FFFFFF")
    def bf():    return Font(name="Arial", size=9)

    # Title row
    ws.merge_cells("A1:S1")
    t           = ws["A1"]
    t.value     = f"Gibson Bank Lending Signal — Results  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    t.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t.fill      = fill(NAVY)
    t.alignment = center
    ws.row_dimensions[1].height = 20

    # Column headers
    cols = [
        ("#",              5),  ("Ticker",       9),  ("Company",      28),
        ("Sector",        20),  ("Period",       12),  ("Filed",        12),
        ("Has Debt?",     10),  ("Lender",       28),  ("Lender Type",  22),
        ("Facility $M",   12),  ("Drawn $M",     11),  ("Rate",         12),
        ("Maturity",      10),  ("Covenants?",   11),  ("Going Concern?",13),
        ("Scrapability",  13),  ("Note Location",20),  ("Typology",     10),
        ("Key Finding",   55),
    ]
    for i, (name, width) in enumerate(cols, 1):
        c           = ws.cell(row=2, column=i, value=name)
        c.font      = hf()
        c.fill      = fill(NAVY)
        c.alignment = center
        c.border    = bdr
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 28

    ticker_info = {t["ticker"]: t for t in TICKERS}

    for row_i, res in enumerate(results, 3):
        ticker      = res.get("ticker", "")
        info        = ticker_info.get(ticker, {})
        ext         = res.get("extracted", {})
        filing      = res.get("filing", {})
        has_debt    = ext.get("has_debt")
        lender_type = ext.get("lender_type", "Unknown")
        going_c     = ext.get("going_concern", False)

        row = [
            row_i - 2,
            ticker,
            info.get("name", ""),
            info.get("sector", ""),
            filing.get("period", ""),
            filing.get("filing_date", ""),
            "Yes" if has_debt else ("No" if has_debt is False else "?"),
            ext.get("lender_name", ""),
            lender_type,
            ext.get("facility_amount_m", ""),
            ext.get("amount_drawn_m", ""),
            ext.get("interest_rate", ""),
            ext.get("maturity", ""),
            "Yes" if ext.get("has_covenants") else ("No" if ext.get("has_covenants") is False else "?"),
            "YES ⚠️" if going_c else "No",
            ext.get("scrapability", ""),
            ext.get("filing_note_location", ""),
            ext.get("typology", ""),
            ext.get("key_finding", res.get("error", "")),
        ]

        for col_i, val in enumerate(row, 1):
            c           = ws.cell(row=row_i, column=col_i, value=val)
            c.font      = bf()
            c.border    = bdr
            c.alignment = left

            if col_i == 7:
                if val == "Yes":  c.fill = fill("FFEB9C")
                elif val == "No": c.fill = fill("C6EFCE")

            if col_i == 9:
                c.fill = fill(LENDER_COLORS.get(lender_type, "F2F2F2"))

            if col_i == 15 and going_c:
                c.fill = fill("FFC7CE")
                c.font = Font(name="Arial", size=9, bold=True, color="9C0006")

        ws.row_dimensions[row_i].height = 42

    wb.save(output_path)
    print(f"\n✅ Excel saved → {output_path}")


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def run():
    print("=" * 55)
    print("  Gibson Bank Lending Signal Pipeline")
    print(f"  {len(TICKERS)} companies to process")
    print("=" * 55)

    results = []
    failed  = []

    for i, info in enumerate(TICKERS, 1):
        ticker = info["ticker"]
        print(f"\n[{i}/{len(TICKERS)}] {ticker} — {info['name']}")

        print("  → Fetching 10-K from EDGAR...")
        filing = get_debt_section(ticker)

        if filing.get("error") and not filing.get("debt_note"):
            print(f"  ⚠️  {filing['error']}")
            failed.append(ticker)

        print("  → Asking Claude to extract debt info...")
        extracted = extract_with_claude(filing)

        results.append({
            "ticker":    ticker,
            "filing":    filing,
            "extracted": extracted,
            "error":     filing.get("error"),
        })

        print(f"  ✓ has_debt={extracted.get('has_debt')} | "
              f"type={extracted.get('lender_type')} | "
              f"typology={extracted.get('typology')}")

        if extracted.get("going_concern"):
            print("  ⚠️  GOING CONCERN flagged!")

        # Save progress after every company
        with open("output/progress.json", "w") as f:
            json.dump(results, f, indent=2, default=str)

        if i < len(TICKERS):
            time.sleep(1.5)

    # Build Excel
    print(f"\n{'='*55}")
    print("  Building Excel output...")
    out = f"output/gibson_results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    build_excel(results, out)

    # Summary
    done    = [r for r in results if r["extracted"].get("has_debt") is not None]
    banks   = [r for r in done if "Traditional commercial bank" in r["extracted"].get("lender_type", "")]
    no_debt = [r for r in done if r["extracted"].get("has_debt") is False]
    going_c = [r for r in done if r["extracted"].get("going_concern")]

    print(f"\n{'='*55}")
    print(f"  DONE!")
    print(f"  Total processed:         {len(results)}")
    print(f"  Traditional bank:        {len(banks)}")
    print(f"  No debt (equity-only):   {len(no_debt)}")
    print(f"  Going concern flagged:   {len(going_c)}")
    print(f"  Needs manual review:     {len(failed)}")
    print(f"{'='*55}")

    if failed:
        print(f"\n  Manual review needed: {', '.join(failed)}")


if __name__ == "__main__":
    run()